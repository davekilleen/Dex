#!/usr/bin/env python3
"""
Dex Customer Intelligence — Monthly Report Generator

Runs headless (no MCP required). Connects directly to Salesforce using
saved OAuth tokens, generates a markdown report, and saves it to the vault.

Usage:
    python3 generate-report.py [--days N] [--output PATH]

    --days N          Look-back window for new assets (default: 30)
    --months N        Look-ahead for lease expirations (default: 12)
    --output PATH     Override output file path
    --stdout          Print report to stdout instead of saving

Automatically run monthly via the launchd agent. Also called from the
/customer-intel report mode in Dex.
"""

import json
import os
import sys
import argparse
import re
from datetime import datetime, date, timedelta
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "lib"))
import sflib

# ── Config ────────────────────────────────────────────────────────────────────

VAULT_PATH = os.environ.get("VAULT_PATH", "")
VAULT_ROOT = Path(VAULT_PATH) if VAULT_PATH else Path(__file__).resolve().parent.parent.parent
SF_DATA_DIR = VAULT_ROOT / ".scripts" / "salesforce-data"
EDA_DATA_DIR = VAULT_ROOT / ".scripts" / "customer-intel" / "eda-data"
CLIENT_ID, CLIENT_SECRET, _OWNER = sflib.resolve_creds(VAULT_ROOT)
REPORT_STATES = {"PA"}

LEGAL_SUFFIXES = {
    "inc", "incorporated", "corp", "corporation", "co", "company",
    "llc", "ltd", "lp", "pllc", "the",
}

US_STATE_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
    "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE",
    "DISTRICT OF COLUMBIA": "DC", "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI",
    "IDAHO": "ID", "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA",
    "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME",
    "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE",
    "NEVADA": "NV", "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM",
    "NEW YORK": "NY", "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH",
    "OKLAHOMA": "OK", "OREGON": "OR", "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI",
    "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD", "TENNESSEE": "TN", "TEXAS": "TX",
    "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA", "WASHINGTON": "WA",
    "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}


# ── Salesforce Auth (delegated to sflib) ──────────────────────────────────────

def get_valid_tokens():
    return sflib.load_tokens()


def sf_query(tokens, soql):
    return sflib.query(tokens, soql)


# ── Owned Account Scope ───────────────────────────────────────────────────────

def normalize_account_name(name):
    if not name:
        return ""
    name = str(name).lower()
    name = re.sub(r"[^a-z0-9 ]", " ", name)
    tokens = [t for t in name.split() if t not in LEGAL_SUFFIXES]
    return re.sub(r"\s+", " ", " ".join(tokens)).strip()


def normalize_state(value):
    if not value:
        return ""
    state = re.sub(r"[^A-Za-z ]", "", str(value)).strip().upper()
    if len(state) == 2:
        return state
    return US_STATE_ABBR.get(state, state)


def normalize_zip(value):
    if not value:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits[:5] if len(digits) >= 5 else ""


def states_for_record(record):
    return {
        state for state in {
            normalize_state(record.get("billing_state")),
            normalize_state(record.get("shipping_state")),
        } if state
    }


def zips_for_record(record):
    return {
        zip_code for zip_code in {
            normalize_zip(record.get("billing_zip")),
            normalize_zip(record.get("shipping_zip")),
        } if zip_code
    }


def city_for_record(record):
    return (
        record.get("billing_city")
        or record.get("shipping_city")
        or "—"
    )


def filter_assets_by_report_state(assets):
    return [
        asset for asset in assets
        if states_for_record(asset) & REPORT_STATES
    ]


def add_months(d, months):
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    days_in_month = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
                     31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(d.day, days_in_month[month - 1])
    return date(year, month, day)


def parse_excel_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(str(value)[:10], fmt).date()
        except Exception:
            pass
    return None


def normalize_buy_ids(value):
    if not value:
        return set()
    pieces = re.split(r"[,;|/\s]+", str(value).upper())
    return {re.sub(r"[^A-Z0-9]", "", p) for p in pieces if re.sub(r"[^A-Z0-9]", "", p)}


def load_owned_account_scope():
    accounts_path = SF_DATA_DIR / "accounts.json"
    if not accounts_path.exists():
        raise FileNotFoundError(
            f"Owned account cache not found: {accounts_path}. "
            "Run .scripts/sf-pull-sync.py --group full before generating this report."
        )

    accounts = json.loads(accounts_path.read_text(encoding="utf-8"))
    by_id = {}
    by_name = {}
    by_buy_id = {}
    duplicate_buy_ids = set()

    for account in accounts:
        account_ref = {
            "id": account.get("Id", ""),
            "name": account.get("Name", ""),
            "billing_city": account.get("BillingCity") or "",
            "billing_state": account.get("BillingState") or "",
            "billing_zip": account.get("BillingPostalCode") or "",
            "shipping_city": account.get("ShippingCity") or "",
            "shipping_state": account.get("ShippingState") or "",
            "shipping_zip": account.get("ShippingPostalCode") or "",
        }
        if account_ref["id"]:
            by_id[account_ref["id"]] = account_ref

        normalized_name = normalize_account_name(account_ref["name"])
        if normalized_name:
            by_name.setdefault(normalized_name, []).append(account_ref)

        for buy_id in normalize_buy_ids(account.get("UCC_BuyID__c")):
            if buy_id in by_buy_id:
                duplicate_buy_ids.add(buy_id)
            else:
                by_buy_id[buy_id] = account_ref

    for buy_id in duplicate_buy_ids:
        by_buy_id.pop(buy_id, None)

    return {
        "by_id": by_id,
        "by_name": by_name,
        "by_buy_id": by_buy_id,
        "count": len(accounts),
        "duplicate_buy_id_count": len(duplicate_buy_ids),
    }


def match_owned_account(asset, scope):
    account = scope["by_id"].get(asset.get("account_id"))
    if account:
        return account, "Account ID", 100

    buy_ids = normalize_buy_ids(asset.get("ucc_buy_id")) | normalize_buy_ids(asset.get("account_ucc_buy_id"))
    for buy_id in buy_ids:
        account = scope["by_buy_id"].get(buy_id)
        if account:
            return account, "UCC Buy ID", 99

    candidates = scope["by_name"].get(normalize_account_name(asset.get("account")), [])
    if candidates:
        asset_states = states_for_record(asset)
        asset_zips = zips_for_record(asset)

        state_matches = [c for c in candidates if asset_states and states_for_record(c) & asset_states]
        zip_matches = [c for c in candidates if asset_zips and zips_for_record(c) & asset_zips]

        if len(state_matches) == 1:
            return state_matches[0], "Name + State", 90
        if len(zip_matches) == 1:
            return zip_matches[0], "Name + ZIP", 95

    return None, None, None


def unmatched_reason(asset, scope):
    reasons = []

    if asset.get("account_id") and asset.get("account_id") not in scope["by_id"]:
        reasons.append("source Account ID not Chris-owned")
    elif not asset.get("account_id"):
        reasons.append("no source Account ID")

    buy_ids = normalize_buy_ids(asset.get("ucc_buy_id")) | normalize_buy_ids(asset.get("account_ucc_buy_id"))
    if buy_ids and not any(buy_id in scope["by_buy_id"] for buy_id in buy_ids):
        reasons.append("UCC Buy ID not in Chris-owned account cache")
    elif not buy_ids:
        reasons.append("no UCC Buy ID")

    normalized_name = normalize_account_name(asset.get("account"))
    candidates = scope["by_name"].get(normalized_name, [])
    if not normalized_name:
        reasons.append("no source account name")
    elif not candidates:
        reasons.append("no exact normalized account-name match")
    else:
        asset_states = states_for_record(asset)
        asset_zips = zips_for_record(asset)
        state_matches = [c for c in candidates if asset_states and states_for_record(c) & asset_states]
        zip_matches = [c for c in candidates if asset_zips and zips_for_record(c) & asset_zips]
        if not state_matches and not zip_matches:
            reasons.append("exact name exists but state/ZIP did not validate")
        elif len(state_matches) > 1 or len(zip_matches) > 1:
            reasons.append("exact name + address match is ambiguous")

    return "; ".join(reasons)


def scope_assets_to_owned_accounts(assets, scope):
    scoped = []
    for asset in assets:
        account, match_method, confidence = match_owned_account(asset, scope)
        if not account:
            continue

        scoped_asset = dict(asset)
        scoped_asset["source_account"] = asset["account"]
        scoped_asset["source_account_id"] = asset["account_id"]
        scoped_asset["matched_account"] = account["name"]
        scoped_asset["matched_account_id"] = account["id"]
        scoped_asset["match_method"] = match_method
        scoped_asset["match_confidence"] = confidence
        scoped_asset["account"] = account["name"] or asset["account"]
        scoped_asset["account_id"] = account["id"] or asset["account_id"]
        scoped.append(scoped_asset)
    return scoped


def unmatched_assets(assets, scope):
    unmatched = []
    for asset in assets:
        account, _, _ = match_owned_account(asset, scope)
        if account:
            continue
        row = dict(asset)
        row["unmatched_reason"] = unmatched_reason(asset, scope)
        unmatched.append(row)
    return unmatched


# ── Asset Parsing ─────────────────────────────────────────────────────────────

def expiry_info(usage_end_str):
    if not usage_end_str:
        return None, None
    try:
        end_date = datetime.strptime(usage_end_str[:10], "%Y-%m-%d").date()
        days = (end_date - date.today()).days
        if days <= 0:
            urgency = "LAPSED"
        elif days <= 90:
            urgency = "CRITICAL"
        elif days <= 180:
            urgency = "HIGH"
        elif days <= 365:
            urgency = "MEDIUM"
        else:
            urgency = "LOW"
        return days, urgency
    except Exception:
        return None, None


def urgency_icon(urgency):
    return {"CRITICAL": "🔴", "HIGH": "🟡", "MEDIUM": "🟠", "LOW": "🟢", "LAPSED": "⚫"}.get(urgency or "", "—")


def parse_record(r):
    days, urgency = expiry_info(r.get("UsageEndDate"))
    return {
        "id": r["Id"],
        "name": r.get("Name", ""),
        "machine_type": r.get("Machine_Type_New__c") or "—",
        "model": r.get("ModelName__c") or "—",
        "builder": r.get("Builder__c") or r.get("UCC_Vendor__c") or "—",
        "serial": r.get("SerialNumber") or "—",
        "sale_or_lease": r.get("Sale_or_Lease__c") or "—",
        "install_date": (r.get("InstallDate") or "")[:10] or "—",
        "usage_end_date": (r.get("UsageEndDate") or "")[:10] or "—",
        "days_to_expiry": days,
        "urgency": urgency,
        "is_competitor": r.get("IsCompetitorProduct", False),
        "account": (r.get("Account") or {}).get("Name") or "Unknown",
        "account_id": (r.get("Account") or {}).get("Id") or "",
        "billing_city": (r.get("Account") or {}).get("BillingCity") or "",
        "billing_state": (r.get("Account") or {}).get("BillingState") or "",
        "billing_zip": (r.get("Account") or {}).get("BillingPostalCode") or "",
        "shipping_city": (r.get("Account") or {}).get("ShippingCity") or "",
        "shipping_state": (r.get("Account") or {}).get("ShippingState") or "",
        "shipping_zip": (r.get("Account") or {}).get("ShippingPostalCode") or "",
        "ucc_buy_id": r.get("UCC_BuyID__c") or "",
        "account_ucc_buy_id": (r.get("Account") or {}).get("UCC_BuyID__c") or "",
        "created_date": (r.get("CreatedDate") or "")[:10] or "",
    }


# ── Salesforce Queries ────────────────────────────────────────────────────────

def get_expiring_assets(tokens, months=12):
    future = (date.today() + timedelta(days=months * 30)).strftime("%Y-%m-%d")
    soql = f"""
        SELECT Id, Name, Machine_Type_New__c, ModelName__c, Builder__c, UCC_Vendor__c,
               Sale_or_Lease__c, UsageEndDate, Status, IsCompetitorProduct, UCC_BuyID__c,
               Account.Name, Account.Id, Account.UCC_BuyID__c,
               Account.BillingCity, Account.BillingState, Account.BillingPostalCode,
               Account.ShippingCity, Account.ShippingState, Account.ShippingPostalCode
        FROM Asset
        WHERE UsageEndDate != null
          AND UsageEndDate >= TODAY
          AND UsageEndDate <= {future}
        ORDER BY UsageEndDate ASC
        LIMIT 500
    """
    result = sf_query(tokens, soql)
    return [parse_record(r) for r in result.get("records", [])]


def get_new_assets(tokens, days=30):
    soql = f"""
        SELECT Id, Name, Machine_Type_New__c, ModelName__c, Builder__c, UCC_Vendor__c,
               Sale_or_Lease__c, InstallDate, UsageEndDate, Status, IsCompetitorProduct,
               UCC_BuyID__c, Account.Name, Account.Id, Account.UCC_BuyID__c,
               Account.BillingCity, Account.BillingState, Account.BillingPostalCode,
               Account.ShippingCity, Account.ShippingState, Account.ShippingPostalCode, CreatedDate
        FROM Asset
        WHERE CreatedDate >= LAST_N_DAYS:{days}
        ORDER BY CreatedDate DESC
        LIMIT 500
    """
    result = sf_query(tokens, soql)
    return [parse_record(r) for r in result.get("records", [])]


def latest_eda_watch_file():
    files = sorted(
        EDA_DATA_DIR.glob("EDA_Export_FullFile_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


def load_eda_lease_watch_assets():
    path = latest_eda_watch_file()
    if not path:
        return [], None

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("WARN: openpyxl not installed; skipping EDA 54-month lease watch XLSX.", file=sys.stderr)
        return [], path

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip().upper() for h in next(rows, [])]
    assets = []

    for idx, row in enumerate(rows, start=2):
        rec = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        company = str(rec.get("BUYCOMP1") or "").strip()
        if not company:
            continue

        filing_date = parse_excel_date(rec.get("UCCDATE"))
        estimated_end = add_months(filing_date, 60) if filing_date else None
        usage_end_date = estimated_end.isoformat() if estimated_end else ""
        days, urgency = expiry_info(usage_end_date)
        ucc_id = str(rec.get("UCCID") or "").strip()
        unit = str(rec.get("EQTUNIT") or "").strip()
        status = str(rec.get("UCCSTATUS") or "").strip().upper()

        assets.append({
            "id": f"eda-watch:{ucc_id}:{unit or idx}",
            "name": " - ".join(p for p in [str(rec.get("EQTMAN") or "").strip(),
                                            str(rec.get("EQTMODEL") or "").strip()] if p) or "EDA Watch Record",
            "machine_type": str(rec.get("EQTDESC") or "").strip() or "—",
            "model": str(rec.get("EQTMODEL") or "").strip() or "—",
            "builder": str(rec.get("EQTMAN") or rec.get("SPCOMP") or "").strip() or "—",
            "serial": str(rec.get("EQTSN") or "").strip() or "—",
            "sale_or_lease": "Lease/Finance Watch" if status in {"LEASE", "SALE"} else (status or "—"),
            "install_date": filing_date.isoformat() if filing_date else "—",
            "usage_end_date": usage_end_date or "—",
            "days_to_expiry": days,
            "urgency": urgency,
            "is_competitor": False,
            "account": company,
            "account_id": "",
            "billing_city": str(rec.get("BUYCITY") or "").strip(),
            "billing_state": str(rec.get("BUYSTATE") or "").strip(),
            "billing_zip": str(rec.get("BUYZIP") or "").strip(),
            "shipping_city": "",
            "shipping_state": "",
            "shipping_zip": "",
            "ucc_buy_id": str(rec.get("BUYID") or "").strip(),
            "account_ucc_buy_id": "",
            "created_date": "",
            "source": "EDA 54-Month Lease Watch",
            "lender": str(rec.get("SPCOMP") or "").strip() or "—",
            "ucc_status": status or "—",
        })

    return assets, path


def filter_expiring_window(assets, months):
    future = date.today() + timedelta(days=months * 30)
    filtered = []
    for asset in assets:
        days = asset.get("days_to_expiry")
        if days is None or days < 0:
            continue
        end = parse_excel_date(asset.get("usage_end_date"))
        if end and end <= future:
            filtered.append(asset)
    return filtered


# ── Report Generation ─────────────────────────────────────────────────────────

def md_table(headers, rows, alignments=None):
    if not rows:
        return "*No records*\n"
    col_widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(cell)))
    sep = "| " + " | ".join("-" * w for w in col_widths) + " |"
    header_row = "| " + " | ".join(h.ljust(col_widths[i]) for i, h in enumerate(headers)) + " |"
    data_rows = ["| " + " | ".join(str(cell).ljust(col_widths[i]) for i, cell in enumerate(row)) + " |" for row in rows]
    return "\n".join([header_row, sep] + data_rows) + "\n"


def build_report(expiring_assets, new_assets, unmatched_expiring, unmatched_new, days_back, months_ahead, generated_at):
    today_str = date.today().strftime("%B %d, %Y")
    month_str = date.today().strftime("%B %Y")

    critical = [a for a in expiring_assets if a["urgency"] == "CRITICAL"]
    high = [a for a in expiring_assets if a["urgency"] == "HIGH"]
    medium = [a for a in expiring_assets if a["urgency"] == "MEDIUM"]

    our_new = [a for a in new_assets if not a["is_competitor"]]
    comp_new = [a for a in new_assets if a["is_competitor"]]
    new_accounts = list({a["account_id"]: a["account"] for a in new_assets if a["account_id"]}.values())

    lines = []
    lines.append(f"# Customer Intelligence Report — {month_str}")
    lines.append(
        f"*Generated: {generated_at} | Source: Salesforce Assets (EDA/UCC-1) | "
        "Scope: PA-only, Chris-owned accounts matched by Account ID, UCC Buy ID, or exact name + address*"
    )
    lines.append("")

    # Summary box
    lines.append("## Summary")
    lines.append("")
    lines.append(f"| Metric | Count |")
    lines.append(f"|--------|-------|")
    lines.append(f"| 🔴 Leases expiring 0–90 days (CRITICAL) | **{len(critical)}** |")
    lines.append(f"| 🟡 Leases expiring 90–180 days (HIGH) | **{len(high)}** |")
    lines.append(f"| 🟠 Leases expiring 180–365 days (MEDIUM) | **{len(medium)}** |")
    lines.append(f"| New equipment records (last {days_back} days) | {len(new_assets)} |")
    lines.append(f"| New accounts with records | {len(new_accounts)} |")
    lines.append(f"| Competitor equipment added | {len(comp_new)} |")
    lines.append(f"| Unmatched source records excluded | {len(unmatched_expiring) + len(unmatched_new)} |")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Expiration sections
    lines.append("## Lease Expirations")
    lines.append("")

    if critical:
        lines.append("### 🔴 CRITICAL — Expiring in 0–90 Days (Act Now)")
        lines.append("")
        rows = [[a["account"], city_for_record(a), a.get("source", "Salesforce Asset"), a["machine_type"], a["model"], a["builder"],
                 a.get("lender", "—"),
                 a["usage_end_date"], f"{a['days_to_expiry']} days",
                 a["match_method"], a["match_confidence"]] for a in critical]
        lines.append(md_table(
            ["Account", "City", "Source", "Machine Type", "Model", "Builder", "Lender", "Lease Ends", "Days Left", "Match", "Conf"],
            rows
        ))
    else:
        lines.append("### 🔴 CRITICAL — Expiring in 0–90 Days")
        lines.append("")
        lines.append("*No critical expirations this period.*")
        lines.append("")

    if high:
        lines.append("### 🟡 HIGH — Expiring in 90–180 Days (Schedule Outreach)")
        lines.append("")
        rows = [[a["account"], city_for_record(a), a.get("source", "Salesforce Asset"), a["machine_type"], a["model"], a["builder"],
                 a.get("lender", "—"),
                 a["usage_end_date"], f"{a['days_to_expiry']} days",
                 a["match_method"], a["match_confidence"]] for a in high]
        lines.append(md_table(
            ["Account", "City", "Source", "Machine Type", "Model", "Builder", "Lender", "Lease Ends", "Days Left", "Match", "Conf"],
            rows
        ))
    else:
        lines.append("### 🟡 HIGH — Expiring in 90–180 Days")
        lines.append("")
        lines.append("*No high-priority expirations this period.*")
        lines.append("")

    if medium:
        lines.append("### 🟠 MEDIUM — Expiring in 180–365 Days (Calendar Reminder)")
        lines.append("")
        rows = [[a["account"], city_for_record(a), a.get("source", "Salesforce Asset"), a["machine_type"], a["model"],
                 a.get("lender", "—"),
                 a["usage_end_date"], f"{a['days_to_expiry']} days",
                 a["match_method"], a["match_confidence"]] for a in medium]
        lines.append(md_table(
            ["Account", "City", "Source", "Machine Type", "Model", "Lender", "Lease Ends", "Days Left", "Match", "Conf"],
            rows
        ))
    else:
        lines.append("### 🟠 MEDIUM — Expiring in 180–365 Days")
        lines.append("")
        lines.append("*No medium-priority expirations this period.*")
        lines.append("")

    lines.append("---")
    lines.append("")

    # New activity section
    lines.append(f"## New Activity (Last {days_back} Days)")
    lines.append("")

    if our_new:
        lines.append(f"### New Equipment Records — Our Machines ({len(our_new)})")
        lines.append("")
        rows = [[a["account"], city_for_record(a), a["machine_type"], a["model"], a["builder"],
                 a["sale_or_lease"], a["install_date"], a["usage_end_date"],
                 a["match_method"], a["match_confidence"]] for a in our_new]
        lines.append(md_table(
            ["Account", "City", "Machine Type", "Model", "Builder", "S/L", "Install Date", "Lease End", "Match", "Conf"],
            rows
        ))
    else:
        lines.append(f"### New Equipment Records — Our Machines")
        lines.append("")
        lines.append("*No new equipment records this period.*")
        lines.append("")

    if comp_new:
        lines.append(f"### New Competitor Equipment ({len(comp_new)})")
        lines.append("")
        rows = [[a["account"], city_for_record(a), a["machine_type"], a["model"], a["builder"],
                 a["install_date"], a["usage_end_date"],
                 a["match_method"], a["match_confidence"]] for a in comp_new]
        lines.append(md_table(
            ["Account", "City", "Machine Type", "Model", "Competitor", "Install Date", "Lease End", "Match", "Conf"],
            rows
        ))
    else:
        lines.append("### New Competitor Equipment")
        lines.append("")
        lines.append("*No new competitor equipment recorded this period.*")
        lines.append("")

    if new_accounts:
        lines.append(f"### New Accounts with Equipment Records ({len(new_accounts)})")
        lines.append("")
        for acct in sorted(new_accounts):
            lines.append(f"- {acct}")
        lines.append("")

    if unmatched_new:
        lines.append(f"### Unmatched New Source Records — Not Included ({len(unmatched_new)})")
        lines.append("")
        rows = [[
            a["account"],
            a.get("source", "Salesforce Asset"),
            city_for_record(a),
            a.get("billing_state") or a.get("shipping_state") or "—",
            a.get("billing_zip") or a.get("shipping_zip") or "—",
            a.get("ucc_buy_id") or a.get("account_ucc_buy_id") or "—",
            a["machine_type"],
            a["model"],
            a["builder"],
            a["unmatched_reason"],
        ] for a in unmatched_new]
        lines.append(md_table(
            ["Source Account", "Source", "City", "State", "ZIP", "UCC Buy ID", "Machine Type", "Model", "Builder", "Reason"],
            rows
        ))

    if unmatched_expiring:
        lines.append(f"### Unmatched Expiration Source Records — Not Included ({len(unmatched_expiring)})")
        lines.append("")
        rows = [[
            a["account"],
            a.get("source", "Salesforce Asset"),
            city_for_record(a),
            a.get("billing_state") or a.get("shipping_state") or "—",
            a.get("billing_zip") or a.get("shipping_zip") or "—",
            a.get("ucc_buy_id") or a.get("account_ucc_buy_id") or "—",
            a["machine_type"],
            a["model"],
            a["usage_end_date"],
            a["unmatched_reason"],
        ] for a in unmatched_expiring]
        lines.append(md_table(
            ["Source Account", "Source", "City", "State", "ZIP", "UCC Buy ID", "Machine Type", "Model", "Lease End", "Reason"],
            rows
        ))

    lines.append("---")
    lines.append("")
    lines.append("## Quick Actions")
    lines.append("")
    lines.append("- Run `/customer-intel alerts` for the interactive lease dashboard")
    lines.append("- Run `/customer-intel [account]` for a full profile on any account above")
    lines.append("- Run `/customer-intel competitive` for the full competitor equipment map")
    lines.append("")
    lines.append("*Report auto-generated by Dex Customer Intelligence automation.*")
    lines.append(f"*Next report: {(date.today().replace(day=1) + timedelta(days=32)).replace(day=1).strftime('%B 1, %Y')}*")

    return "\n".join(lines) + "\n"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Dex Customer Intelligence report")
    parser.add_argument("--days", type=int, default=30, help="Look-back window for new assets (default: 30)")
    parser.add_argument("--months", type=int, default=12, help="Look-ahead for lease expirations (default: 12)")
    parser.add_argument("--output", type=str, default="", help="Override output file path")
    parser.add_argument("--stdout", action="store_true", help="Print to stdout instead of saving")
    args = parser.parse_args()

    tokens = get_valid_tokens()
    if not tokens:
        print("ERROR: No Salesforce tokens found. Run sf_authenticate from Dex first.", file=sys.stderr)
        sys.exit(1)

    try:
        owned_scope = load_owned_account_scope()
    except Exception as e:
        print(f"ERROR: Could not load Chris-owned account scope: {e}", file=sys.stderr)
        sys.exit(1)

    print(
        f"Loaded Chris-owned account scope: {owned_scope['count']} accounts, "
        f"{len(owned_scope['by_buy_id'])} unique UCC Buy IDs",
        file=sys.stderr,
    )
    if owned_scope["duplicate_buy_id_count"]:
        print(
            f"Skipped {owned_scope['duplicate_buy_id_count']} duplicate UCC Buy IDs in account scope",
            file=sys.stderr,
        )

    print("Fetching lease expirations...", file=sys.stderr)
    raw_sf_expiring = get_expiring_assets(tokens, months=args.months)
    raw_eda_watch, eda_watch_path = load_eda_lease_watch_assets()
    raw_eda_expiring = filter_expiring_window(raw_eda_watch, args.months)
    if eda_watch_path:
        print(f"Loaded EDA lease watch: {eda_watch_path} ({len(raw_eda_expiring)} in window)", file=sys.stderr)
    else:
        print("No EDA lease watch XLSX found; lease expirations use Salesforce Assets only.", file=sys.stderr)
    raw_expiring = filter_assets_by_report_state(raw_sf_expiring + raw_eda_expiring)
    expiring = scope_assets_to_owned_accounts(raw_expiring, owned_scope)
    unmatched_expiring = unmatched_assets(raw_expiring, owned_scope)
    print(f"Scoped lease expirations: {len(raw_expiring)} -> {len(expiring)}", file=sys.stderr)

    print(f"Fetching new assets (last {args.days} days)...", file=sys.stderr)
    raw_new = filter_assets_by_report_state(get_new_assets(tokens, days=args.days))
    new = scope_assets_to_owned_accounts(raw_new, owned_scope)
    unmatched_new = unmatched_assets(raw_new, owned_scope)
    print(f"Scoped new assets: {len(raw_new)} -> {len(new)}", file=sys.stderr)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    report = build_report(expiring, new, unmatched_expiring, unmatched_new, args.days, args.months, generated_at)

    if args.stdout:
        print(report)
        return

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    elif VAULT_PATH:
        report_dir = Path(VAULT_PATH) / "Inbox" / "Reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        filename = f"Customer_Intel_{date.today().strftime('%Y-%m')}.md"
        output_path = report_dir / filename
    else:
        # Fallback: write next to this script
        script_dir = Path(__file__).parent
        output_path = script_dir / f"customer-intel-{date.today().strftime('%Y-%m')}.md"

    output_path.write_text(report, encoding="utf-8")
    print(f"Report saved: {output_path}", file=sys.stderr)
    print(str(output_path))  # stdout = path, for the installer to use


if __name__ == "__main__":
    main()
